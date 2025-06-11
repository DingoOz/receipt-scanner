import json
import csv
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime
from decimal import Decimal

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference

from ..processing.data_extractor import ReceiptData, ReceiptItem
from ..utils.config import ExportConfig


class SpreadsheetExporter:
    """Exports receipt data to various spreadsheet formats."""
    
    def __init__(self, config: ExportConfig):
        """
        Initialize spreadsheet exporter.
        
        Args:
            config: Export configuration
        """
        self.config = config
        self.logger = logging.getLogger(__name__)
        
        # Create output directory
        self.output_dir = Path(config.output_directory)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_receipts(self, ocr_results: List[Dict[str, Any]], source_name: str) -> Dict[str, Any]:
        """
        Export receipt data to configured format.
        
        Args:
            ocr_results: List of OCR processing results
            source_name: Name of the source (folder/album)
            
        Returns:
            Dict with export results
        """
        try:
            self.logger.info(f"Exporting {len(ocr_results)} receipts to {self.config.output_format}")
            
            # Filter successful receipts
            valid_receipts = [
                result for result in ocr_results 
                if result.get('success') and result.get('receipt_data')
            ]
            
            if not valid_receipts:
                return {
                    'success': False,
                    'error': 'No valid receipts to export',
                    'exported_files': []
                }
            
            # Generate timestamp for filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_filename = f"receipts_{source_name}_{timestamp}".replace(' ', '_')
            
            exported_files = []
            
            if self.config.output_format.lower() == 'csv':
                csv_file = self._export_csv(valid_receipts, base_filename)
                exported_files.append(csv_file)
                
            elif self.config.output_format.lower() == 'xlsx':
                xlsx_file = self._export_excel(valid_receipts, base_filename)
                exported_files.append(xlsx_file)
                
            elif self.config.output_format.lower() == 'json':
                json_file = self._export_json(valid_receipts, base_filename)
                exported_files.append(json_file)
                
            else:
                # Export all formats
                csv_file = self._export_csv(valid_receipts, base_filename)
                xlsx_file = self._export_excel(valid_receipts, base_filename)
                json_file = self._export_json(valid_receipts, base_filename)
                exported_files.extend([csv_file, xlsx_file, json_file])
            
            return {
                'success': True,
                'exported_files': exported_files,
                'receipts_exported': len(valid_receipts),
                'output_directory': str(self.output_dir)
            }
            
        except Exception as e:
            self.logger.error(f"Export failed: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'exported_files': []
            }
    
    def _export_csv(self, receipts: List[Dict[str, Any]], base_filename: str) -> str:
        """Export receipts to CSV format."""
        csv_path = self.output_dir / f"{base_filename}.csv"
        
        # Prepare data for CSV
        csv_data = []
        
        for receipt_result in receipts:
            receipt_data = receipt_result['receipt_data']
            base_row = {
                'file_name': receipt_result.get('file_name', 'Unknown'),
                'merchant_name': receipt_data.get('merchant_name', ''),
                'merchant_address': receipt_data.get('merchant_address', ''),
                'merchant_phone': receipt_data.get('merchant_phone', ''),
                'date': receipt_data.get('date', ''),
                'time': receipt_data.get('time', ''),
                'subtotal': receipt_data.get('subtotal', ''),
                'tax_amount': receipt_data.get('tax_amount', ''),
                'tip_amount': receipt_data.get('tip_amount', ''),
                'total_amount': receipt_data.get('total_amount', ''),
                'payment_method': receipt_data.get('payment_method', ''),
                'card_last_four': receipt_data.get('card_last_four', ''),
                'receipt_number': receipt_data.get('receipt_number', ''),
                'ocr_method': receipt_result.get('ocr_method', ''),
                'ocr_confidence': receipt_result.get('ocr_confidence', ''),
                'validation_confidence': receipt_data.get('confidence_score', '')
            }
            
            # Add confidence scores if enabled
            if self.config.include_confidence_scores:
                validation = receipt_result.get('validation', {})
                base_row.update({
                    'validation_is_valid': validation.get('is_valid', ''),
                    'validation_score': validation.get('confidence_score', '')
                })
            
            # Add raw text if enabled
            if self.config.include_raw_text:
                base_row['raw_text'] = receipt_result.get('raw_text', '')
            
            # If no items, add one row with receipt data
            items = receipt_data.get('items', [])
            if not items:
                csv_data.append(base_row)
            else:
                # Add one row per item
                for item in items:
                    item_row = base_row.copy()
                    item_row.update({
                        'item_description': item.get('description', ''),
                        'item_quantity': item.get('quantity', ''),
                        'item_unit_price': item.get('unit_price', ''),
                        'item_total_price': item.get('total_price', ''),
                        'item_confidence': item.get('confidence', '')
                    })
                    csv_data.append(item_row)
        
        # Write CSV
        if csv_data:
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = csv_data[0].keys()
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(csv_data)
        
        self.logger.info(f"CSV exported to: {csv_path}")
        return str(csv_path)
    
    def _export_excel(self, receipts: List[Dict[str, Any]], base_filename: str) -> str:
        """Export receipts to Excel format with formatting and charts."""
        xlsx_path = self.output_dir / f"{base_filename}.xlsx"
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet and create our sheets
        wb.remove(wb.active)
        
        # Create worksheets
        summary_ws = wb.create_sheet("Summary")
        receipts_ws = wb.create_sheet("Receipt Details")
        items_ws = wb.create_sheet("Line Items")
        
        # Export data to each sheet
        self._create_summary_sheet(summary_ws, receipts)
        self._create_receipts_sheet(receipts_ws, receipts)
        self._create_items_sheet(items_ws, receipts)
        
        # Add charts if there's enough data
        if len(receipts) > 1:
            self._add_charts_to_summary(summary_ws, receipts)
        
        # Save workbook
        wb.save(xlsx_path)
        
        self.logger.info(f"Excel exported to: {xlsx_path}")
        return str(xlsx_path)
    
    def _create_summary_sheet(self, ws, receipts: List[Dict[str, Any]]):
        """Create summary worksheet."""
        # Title
        ws['A1'] = 'Receipt Processing Summary'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Summary statistics
        total_receipts = len(receipts)
        total_amount = sum(
            float(r['receipt_data'].get('total_amount', 0)) 
            for r in receipts 
            if r['receipt_data'].get('total_amount')
        )
        
        merchants = set(
            r['receipt_data'].get('merchant_name') 
            for r in receipts 
            if r['receipt_data'].get('merchant_name')
        )
        
        avg_confidence = sum(
            float(r['receipt_data'].get('confidence_score', 0)) 
            for r in receipts
        ) / len(receipts) if receipts else 0
        
        # Statistics table
        stats_data = [
            ['Statistic', 'Value'],
            ['Total Receipts', total_receipts],
            ['Total Amount', f'${total_amount:.2f}'],
            ['Unique Merchants', len(merchants)],
            ['Average Confidence', f'{avg_confidence:.1%}'],
            ['Export Date', datetime.now().strftime(self.config.date_format)]
        ]
        
        for row_idx, row_data in enumerate(stats_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Merchant breakdown
        ws['A10'] = 'Merchant Breakdown'
        ws['A10'].font = Font(size=14, bold=True)
        
        merchant_amounts = {}
        merchant_counts = {}
        
        for receipt in receipts:
            merchant = receipt['receipt_data'].get('merchant_name', 'Unknown')
            amount = float(receipt['receipt_data'].get('total_amount', 0)) if receipt['receipt_data'].get('total_amount') else 0
            
            merchant_amounts[merchant] = merchant_amounts.get(merchant, 0) + amount
            merchant_counts[merchant] = merchant_counts.get(merchant, 0) + 1
        
        # Sort by total amount
        sorted_merchants = sorted(merchant_amounts.items(), key=lambda x: x[1], reverse=True)
        
        merchant_data = [['Merchant', 'Count', 'Total Amount', 'Average']]
        for merchant, total_amount in sorted_merchants[:10]:  # Top 10
            count = merchant_counts[merchant]
            avg_amount = total_amount / count if count > 0 else 0
            merchant_data.append([merchant, count, f'${total_amount:.2f}', f'${avg_amount:.2f}'])
        
        for row_idx, row_data in enumerate(merchant_data, start=12):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 12:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_receipts_sheet(self, ws, receipts: List[Dict[str, Any]]):
        """Create receipts details worksheet."""
        # Headers
        headers = [
            'File Name', 'Merchant', 'Date', 'Time', 'Subtotal', 'Tax', 'Tip', 
            'Total', 'Payment Method', 'Receipt #', 'OCR Method', 'OCR Confidence',
            'Validation Score', 'Items Count'
        ]
        
        if self.config.include_confidence_scores:
            headers.extend(['Validation Details', 'Is Valid'])
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Write data
        for row_idx, receipt in enumerate(receipts, start=2):
            rd = receipt['receipt_data']
            
            row_data = [
                receipt.get('file_name', ''),
                rd.get('merchant_name', ''),
                rd.get('date', ''),
                rd.get('time', ''),
                rd.get('subtotal', ''),
                rd.get('tax_amount', ''),
                rd.get('tip_amount', ''),
                rd.get('total_amount', ''),
                rd.get('payment_method', ''),
                rd.get('receipt_number', ''),
                receipt.get('ocr_method', ''),
                f"{receipt.get('ocr_confidence', 0):.1%}",
                f"{rd.get('confidence_score', 0):.1%}",
                len(rd.get('items', []))
            ]
            
            if self.config.include_confidence_scores:
                validation = receipt.get('validation', {})
                row_data.extend([
                    f"{validation.get('confidence_score', 0):.1%}",
                    'Yes' if validation.get('is_valid', False) else 'No'
                ])
            
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)  # Cap at 50
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_items_sheet(self, ws, receipts: List[Dict[str, Any]]):
        """Create line items worksheet."""
        # Headers
        headers = [
            'Receipt File', 'Merchant', 'Date', 'Item Description', 'Quantity', 
            'Unit Price', 'Total Price', 'Item Confidence'
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Write data
        row_idx = 2
        for receipt in receipts:
            rd = receipt['receipt_data']
            items = rd.get('items', [])
            
            if not items:
                # Add a row even if no items
                row_data = [
                    receipt.get('file_name', ''),
                    rd.get('merchant_name', ''),
                    rd.get('date', ''),
                    '(No items detected)',
                    '', '', '', ''
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                row_idx += 1
            else:
                for item in items:
                    row_data = [
                        receipt.get('file_name', ''),
                        rd.get('merchant_name', ''),
                        rd.get('date', ''),
                        item.get('description', ''),
                        item.get('quantity', ''),
                        item.get('unit_price', ''),
                        item.get('total_price', ''),
                        f"{item.get('confidence', 0):.1%}" if item.get('confidence') else ''
                    ]
                    
                    for col_idx, value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                    row_idx += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)  # Cap at 50
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _add_charts_to_summary(self, ws, receipts: List[Dict[str, Any]]):
        """Add charts to summary sheet."""
        try:
            # Merchant spending chart
            merchant_amounts = {}
            for receipt in receipts:
                merchant = receipt['receipt_data'].get('merchant_name', 'Unknown')
                amount = float(receipt['receipt_data'].get('total_amount', 0)) if receipt['receipt_data'].get('total_amount') else 0
                merchant_amounts[merchant] = merchant_amounts.get(merchant, 0) + amount
            
            # Only add chart if we have data
            if len(merchant_amounts) > 1:
                # Create pie chart for top merchants
                chart = PieChart()
                chart.title = "Spending by Merchant"
                
                # We already have merchant data in the sheet, reference it
                # This is a simplified approach - in a full implementation,
                # you'd want to set up the data references properly
                chart.width = 15
                chart.height = 10
                
                # Position chart
                ws.add_chart(chart, "F3")
                
        except Exception as e:
            self.logger.warning(f"Failed to add charts: {str(e)}")
    
    def _export_json(self, receipts: List[Dict[str, Any]], base_filename: str) -> str:
        """Export receipts to JSON format."""
        json_path = self.output_dir / f"{base_filename}.json"
        
        # Prepare data for JSON
        export_data = {
            'export_info': {
                'timestamp': datetime.now().isoformat(),
                'total_receipts': len(receipts),
                'exporter_version': '1.0',
                'include_confidence_scores': self.config.include_confidence_scores,
                'include_raw_text': self.config.include_raw_text
            },
            'receipts': []
        }
        
        for receipt in receipts:
            receipt_export = {
                'file_info': {
                    'file_name': receipt.get('file_name', ''),
                    'file_id': receipt.get('file_id', ''),
                    'processing_time': receipt.get('processing_time', 0)
                },
                'ocr_info': {
                    'method': receipt.get('ocr_method', ''),
                    'confidence': receipt.get('ocr_confidence', 0)
                },
                'receipt_data': receipt['receipt_data']
            }
            
            # Add validation info if confidence scores enabled
            if self.config.include_confidence_scores and 'validation' in receipt:
                receipt_export['validation'] = receipt['validation']
            
            # Add raw text if enabled
            if self.config.include_raw_text and 'raw_text' in receipt:
                receipt_export['raw_text'] = receipt['raw_text']
            
            # Add quality metrics if available
            if 'quality_metrics' in receipt:
                receipt_export['quality_metrics'] = receipt['quality_metrics']
            
            export_data['receipts'].append(receipt_export)
        
        # Write JSON
        with open(json_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(export_data, jsonfile, indent=2, ensure_ascii=False, default=str)
        
        self.logger.info(f"JSON exported to: {json_path}")
        return str(json_path)
    
    def export_summary_report(self, receipts: List[Dict[str, Any]], source_name: str) -> str:
        """
        Generate a summary report in text format.
        
        Args:
            receipts: List of receipt results
            source_name: Name of the source
            
        Returns:
            Path to generated report
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_path = self.output_dir / f"summary_report_{source_name}_{timestamp}.txt"
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("="*60 + "\n")
            f.write("RECEIPT PROCESSING SUMMARY REPORT\n")
            f.write("="*60 + "\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Source: {source_name}\n")
            f.write("="*60 + "\n\n")
            
            # Overall statistics
            total_receipts = len(receipts)
            valid_receipts = [r for r in receipts if r.get('success')]
            total_amount = sum(
                float(r['receipt_data'].get('total_amount', 0)) 
                for r in valid_receipts 
                if r['receipt_data'].get('total_amount')
            )
            
            f.write("OVERALL STATISTICS:\n")
            f.write("-" * 20 + "\n")
            f.write(f"Total Receipts Processed: {total_receipts}\n")
            f.write(f"Successfully Parsed: {len(valid_receipts)}\n")
            f.write(f"Success Rate: {len(valid_receipts)/total_receipts*100:.1f}%\n")
            f.write(f"Total Amount: ${total_amount:.2f}\n\n")
            
            # Merchant breakdown
            merchants = {}
            for receipt in valid_receipts:
                merchant = receipt['receipt_data'].get('merchant_name', 'Unknown')
                amount = float(receipt['receipt_data'].get('total_amount', 0)) if receipt['receipt_data'].get('total_amount') else 0
                if merchant not in merchants:
                    merchants[merchant] = {'count': 0, 'total': 0}
                merchants[merchant]['count'] += 1
                merchants[merchant]['total'] += amount
            
            f.write("MERCHANT BREAKDOWN:\n")
            f.write("-" * 20 + "\n")
            for merchant, data in sorted(merchants.items(), key=lambda x: x[1]['total'], reverse=True):
                f.write(f"{merchant}: {data['count']} receipts, ${data['total']:.2f}\n")
            f.write("\n")
            
            # OCR method breakdown
            ocr_methods = {}
            for receipt in valid_receipts:
                method = receipt.get('ocr_method', 'unknown')
                ocr_methods[method] = ocr_methods.get(method, 0) + 1
            
            f.write("OCR METHODS USED:\n")
            f.write("-" * 20 + "\n")
            for method, count in ocr_methods.items():
                f.write(f"{method}: {count} receipts\n")
            f.write("\n")
            
            # Individual receipt details
            f.write("INDIVIDUAL RECEIPTS:\n")
            f.write("-" * 20 + "\n")
            for i, receipt in enumerate(valid_receipts, 1):
                rd = receipt['receipt_data']
                f.write(f"{i}. {receipt.get('file_name', 'Unknown')}\n")
                f.write(f"   Merchant: {rd.get('merchant_name', 'Unknown')}\n")
                f.write(f"   Date: {rd.get('date', 'Unknown')}\n")
                f.write(f"   Total: ${rd.get('total_amount', '0.00')}\n")
                f.write(f"   Confidence: {rd.get('confidence_score', 0):.1%}\n")
                f.write(f"   Items: {len(rd.get('items', []))}\n\n")
        
        self.logger.info(f"Summary report generated: {report_path}")
        return str(report_path)